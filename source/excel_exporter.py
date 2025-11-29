"""
Excel exporter for valid and flagged rows.
Supports formatting, multiple sheets, and styling.
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Row, RowStatus
from .exporter import Exporter

logger = logging.getLogger(__name__)


class ExcelExporter(Exporter):
    """Export rows to Excel format."""
    
    def export(self, rows: List[Row], filename: str, **options) -> Path:
        """
        Export rows to Excel.
        
        Args:
            rows: List of Row objects
            filename: Output filename
            **options:
                - status: RowStatus filter (VALID, FLAGGED, None, or 'BOTH')
                - fields: List of fields to include
                - include_metadata: Include _row_id, _status
                - include_errors: Include validation errors
                - sheet_name: Custom sheet name
                - auto_format: Apply formatting (header, widths, etc.)
                - freeze_header: Freeze header row
        
        Returns:
            Path to exported Excel file
        """
        status = options.get('status')
        fields = options.get('fields')
        include_metadata = options.get('include_metadata', False)
        include_errors = options.get('include_errors', True)
        sheet_name = options.get('sheet_name', 'Data')
        auto_format = options.get('auto_format', True)
        freeze_header = options.get('freeze_header', True)
        
        output_path = self.output_dir / filename
        workbook = Workbook()
        
        try:
            # Handle BOTH status (separate sheets)
            if status == 'BOTH':
                workbook.remove(workbook.active)
                
                valid_rows = self._filter_rows(rows, RowStatus.VALID)
                if valid_rows:
                    self._write_sheet(workbook, valid_rows, 'Valid', 
                                    fields, include_metadata, include_errors,
                                    auto_format, freeze_header)
                
                flagged_rows = self._filter_rows(rows, RowStatus.FLAGGED)
                if flagged_rows:
                    self._write_sheet(workbook, flagged_rows, 'Flagged', 
                                    fields, include_metadata, include_errors,
                                    auto_format, freeze_header)
            else:
                filtered_rows = self._filter_rows(rows, status)
                if not filtered_rows:
                    logger.warning(f"No rows to export for status {status}")
                    return None
                
                self._write_sheet(workbook, filtered_rows, sheet_name, 
                                fields, include_metadata, include_errors,
                                auto_format, freeze_header)
            
            workbook.save(output_path)
            self._log_export(filename, len(rows), status)
            return output_path
        
        except Exception as e:
            logger.error(f"Error exporting Excel: {e}")
            raise
    
    def _write_sheet(self, workbook, rows: List[Row], sheet_name: str,
                    fields: Optional[List[str]], include_metadata: bool,
                    include_errors: bool, auto_format: bool, freeze_header: bool):
        """Write rows to a worksheet."""
        ws = workbook.create_sheet(sheet_name)
        
        # Collect fields
        all_fields = set()
        for row in rows:
            if row.canonical_data:
                all_fields.update(row.canonical_data.keys())
        
        field_list = fields or sorted(list(all_fields))
        
        # Build header
        header = field_list.copy()
        if include_metadata:
            header.extend(['_row_id', '_line_number', '_status'])
        if include_errors:
            header.append('_errors')
        
        ws.append(header)
        
        # Format header if enabled
        if auto_format:
            header_fill = PatternFill(start_color="366092", end_color="366092", 
                                     fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write rows
        for row in rows:
            row_dict = self._extract_fields(row, field_list)
            
            if include_metadata:
                row_dict = self._add_metadata(row_dict, row, include_errors)
            
            row_values = [row_dict.get(f, '') for f in header]
            ws.append(row_values)
        
        # Auto-format columns
        if auto_format:
            for idx, field in enumerate(header, 1):
                col_letter = get_column_letter(idx)
                ws.column_dimensions[col_letter].width = min(20, len(field) + 2)
            
            # Alternate row colors
            light_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", 
                                    fill_type="solid")
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
                if row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = light_fill
        
        # Freeze header
        if freeze_header:
            ws.freeze_panes = 'A2'
